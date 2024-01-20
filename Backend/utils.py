from datetime import date, timedelta
from Backend.models import Client,Quotation,Miss_Food_Application,Work_Item,Employee,Travel_Application,Clock_Correction_Application, Work_Overtime_Application, Salary,SalaryDetail,Leave_Param,Leave_Application, Clock,Project_Confirmation,Project_Job_Assign,Project_Employee_Assign
from urllib.parse import parse_qs
from django.forms.models import model_to_dict
from django.conf import settings
from django.http import HttpResponse
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from urllib.parse import quote
import os
from django.db.models import Sum
import math
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill



#
def Check_Permissions(user,name,check_get=False,get_name=""):

    if settings.PASS_TEST_FUNC or user.groups.filter(name__icontains='最高權限').exists() :
        return True
    if check_get=="GET": #請求的模式如果是get，就判斷另一個get_name
        return user.groups.filter(name__icontains=get_name).exists()  or user.groups.filter(name__icontains=name).exists()  

    return user.groups.filter(name__icontains=name).exists()  


#取得當周日期
def get_weekdays(to_week):
    weekdays = []
    today = date.today()
    current_date = today - timedelta(days=today.weekday())
    while current_date.weekday() < to_week:
        weekdays.append(current_date)
        current_date += timedelta(days=1)
    return weekdays 

def get_weekly_clock_data(userid):
    weekdays = get_weekdays(5)
    weekly_clock_data = []

    for weekday in weekdays:        
        clock_records = Clock.objects.filter(clock_date=weekday).filter(employee_id=userid).order_by('clock_time')
        clock_in_records = []
        clock_out_records = []
        for record in clock_records:

            if record.type_of_clock =="2":
                clock_corrections = record.clock_correction.all()

                if len(clock_corrections) :
                    get_approval = clock_corrections[0].Approval
                    if get_approval:
                        if get_approval.current_status =="completed":                    
                            if record.clock_in_or_out:
                                clock_in_records.append(record)   
                            else:         
                                clock_out_records.append(record)   
            else:
                if record.clock_in_or_out:
                    clock_in_records.append(record)   
                else:         
                    clock_out_records.append(record)   

        if len(clock_in_records)==0:
            check_in = ""
        else:
            check_in =clock_in_records[0].clock_time.strftime('%H:%M') 

        if len(clock_out_records)==0:
            check_out = ""
        else:
            check_out = clock_out_records[-1].clock_time.strftime('%H:%M') 


      
        daily_data = {
            'day': weekday.strftime('%m/%d'),
            'checkin': check_in,
            'checkout': check_out
        }
        weekly_clock_data.append(daily_data)
    return weekly_clock_data

def convent_employee(employees):
    employee_ary=[]
    selected_fields = ['id','full_name']
    for get_employee in employees:
        employee_dict = model_to_dict(get_employee,fields=selected_fields)
        
        # del employee_dict["uploaded_files"]
        # if "profile_image" in employee_dict:
        #     if  employee_dict['profile_image']:
        #         employee_dict['profile_image'] = employee_dict["profile_image"].url
        #     else:
        #         employee_dict['profile_image'] = None       
        employee_ary.append(employee_dict)
    return employee_ary


def convent_dict(data):
    data_str = data.decode('utf-8')
    dict_data = parse_qs(data_str)
    print(dict_data)
    print("convent")
    if "csrfmiddlewaretoken" in dict_data:
        del dict_data["csrfmiddlewaretoken"]


    new_dict_data = {}
    for key, value in dict_data.items():
        new_dict_data[key] = value[0]
        process_key =("vehicle","Work_Item_Number_id","work_item_number","work_item_id","inspector","support_employee","work_item","carry_equipments","user_set","completion_report_employee","work_employee","lead_employee","completion_report_employeeS")
        if key in process_key: #處理員工多對多陣列        
            new_dict_data[key] =  [int(num) for num in  value]
        elif  key == "test_items":           
            new_dict_data[key] =[num for num in  value]
        else:            
            match  value[0]:
                case "true":
                    new_dict_data[key] = True               
                case "false":
                    new_dict_data[key] =False
                case _:
                    new_dict_data[key] = value[0]
    print(new_dict_data)
    return new_dict_data


def convent_excel_dict(worksheet,model):
    template_dict={}
    convent_model= None
    print("model: ",model)
    match model:
        case "Employee":
            template_dict= {
                "full_name":"",
                "employee_id":"",
                "id_number":"",
                "gender":"",
                "blood_type":"",
                "departments":"",
                "position":"",
                "start_work_date":"",
                "location_city":"",
            }
            convent_model=Employee
        case "project-confirmation":
            template_dict= {
                "quotation":"",
                "order_id":"",
                "c_a":"",
                "turnover":"",
                "remark":"",
            }
            convent_model=Project_Confirmation
        case "job-assign":
            template_dict= {
                "project_confirmation":"",
                "attendance_date":"",
                "location":"",
                "remark":"",
            }
            convent_model=Project_Job_Assign
        case "employee-assign":
            template_dict= {
                "project_job_assign":"",
                "construction_date":"",
                "completion_date":"",
                "manuscript_return_date":"",
                "remark":"",
            }
            convent_model=Project_Employee_Assign
        case "Work_Item":
            template_dict= {
                "item_id":"",
                "item_name":"",
                "contract_id":"",
                "requisition":"",
                "unit":"",
                "date":"",#額外欄位
                "money":"",#額外欄位
            }
            convent_model=Work_Item
        case "Client": # 客戶公司管理
            template_dict= {
                "client_name":"",
                "client_chinese_name":"",
                "client_english_name":"",
                "client_id":"",
                "tax_id":"",
                "established":"",
                "contact_principal":"",
                "pay_days":"",
                "pay_method":"",
            }
            convent_model=Client
        case "Quotation": # 報價單
            template_dict= {
                "quotation_id":"",
                "project_name":"",
                "quote_date":"",
                "quote_validity_period":"",
                "pay_method":"",
                "internal_content":"",
                "remark":"",
            }
            convent_model=Quotation
        case _:
            return "error"

    convent_ary=[]
    pass_one =True#跳過第一航
    for row in worksheet.iter_rows():
        if pass_one:
            pass_one =False
            continue
            
        new_dict = template_dict.copy()
        for i, key in enumerate(new_dict.keys()):
            print(row[i])
            print(row[i].value)
            new_dict[key] = row[i].value
        
        #防止全空白+入
        all_values_are_none_or_blank = all(value is None or value == '' for value in new_dict.values())
        if not all_values_are_none_or_blank: 
            convent_ary.append(new_dict)

    return convent_ary,convent_model

def get_model_by_name(model_name):
     match model_name:
        case "project_confirmation":
            return Project_Confirmation
        case "job_assign":
            return Project_Job_Assign
        case "Project_Employee_Assign":
            return Project_Employee_Assign
        case "Work_Overtime_Application":
            return Work_Overtime_Application
        case "Leave_Application":
             return Leave_Application
        case "Clock_Correction_Application":
             return Clock_Correction_Application
        case "Travel_Application":
             return Travel_Application
        case "Miss_Food_Application":
             return Miss_Food_Application
        case _:
            return None