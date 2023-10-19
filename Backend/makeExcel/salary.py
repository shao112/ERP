from datetime import date, timedelta
from Backend.models import Employee,Travel_Application,Clock_Correction_Application, Work_Overtime_Application, Salary,SalaryDetail,Leave_Param,Leave_Application, Clock,Project_Confirmation,Project_Job_Assign,Project_Employee_Assign
from urllib.parse import parse_qs
from django.forms.models import model_to_dict
from django.conf import settings
from django.http import HttpResponse
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from urllib.parse import quote
import os
from cn2an import an2cn
from django.db.models import Sum
import math
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill


def salaryFile(get_salary,get_type):
    title = ""
    if get_type:
        title = "薪資條"
    else:
        title = "激勵性獎金"

    get_details = get_salary.details.all()
    file_path = r'media/system_files/salary_template.xlsx'
    miss_obj = 0 #取得勞退項目

    try:
        miss_obj =get_salary.details.filter(name="*勞退公提(6%免扣)").first().adjustment_amount
        print("x")
        print(miss_obj)
    except Exception as e:
       pass

    try:
        workbook = load_workbook(filename=file_path)
    except Exception as e:
        return True,"檔案位置發生問題"

    year,month,user = get_salary.year, get_salary.month,get_salary.user
    full_name,employee_id=user.full_name,user.employee_id,
    departments_name = ""
    if user.departments:
        departments_name = user.departments.department_name

    sheet = workbook.active
    
    deduction_items = get_details.filter(deduction=True,five=get_type) #.exclude(name="*勞退")#扣
    addition_items = get_details.filter(deduction=False,five=get_type)#加


    deduction_sum = deduction_items.aggregate(total=Sum('adjustment_amount'))['total'] or 0
    addition_sum = addition_items.aggregate(total=Sum('adjustment_amount'))['total'] or 0


    deduction_sum -=miss_obj #扣項把勞退扣回去

    difference = addition_sum - deduction_sum

    print(full_name,"add",addition_sum,"cost",deduction_sum,"toto",difference)

    #http://localhost:8000/restful/salaryfile/2023/10/2/1
    try:
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            
            for index, cell_value in enumerate(row, start=1):
                if cell_value == "YM":
                    sheet.cell(row=i, column=index, value=f"{year}年{month}月")
       
                if cell_value == "Title":
                    sheet.cell(row=i, column=index, value=title)
                elif cell_value == "E_ID":
                    sheet.cell(row=i, column=index, value=employee_id)
                elif cell_value == "D_SUM":
                    sheet.cell(row=i, column=index, value=addition_sum)
                elif cell_value == "D_COST":
                    sheet.cell(row=i, column=index, value=deduction_sum)
                elif cell_value == "D_RESULT":
                    sheet.cell(row=i, column=index, value=difference)
                elif cell_value == "DT":
                    sheet.cell(row=i, column=index, value=departments_name)
                elif cell_value == "NAME":
                    sheet.cell(row=i, column=index, value=full_name)
                if cell_value == "ADD_LIST":
                    if  len(addition_items)==0:
                        sheet.cell(row=i, column=index, value="")
                        sheet.cell(row=i, column=index+1, value="")
                    else:
                        for next_index, item in enumerate(addition_items):
                            sheet.cell(row=i+next_index, column=index, value= item.name)
                            sheet.cell(row=i+next_index, column=index+1, value=item.adjustment_amount)
                if cell_value == "COST_LIST":
                    if  len(deduction_items)==0:
                        sheet.cell(row=i, column=index, value="")
                        sheet.cell(row=i, column=index+1, value="")
                    else:
                        for next_index, item in enumerate(deduction_items):
                            if item.name=="*勞退公提(6%免扣)":
                                sheet.cell(row=i+next_index, column=index, value= "勞退公提(6%免扣)")
                                sheet.cell(row=i+next_index, column=index+1, value=item.adjustment_amount)
                            else:
                                sheet.cell(row=i+next_index, column=index, value= item.name)
                                sheet.cell(row=i+next_index, column=index+1, value=item.adjustment_amount)

    except ValueError as e:
        return True,"遇到欄位合併的錯誤"
    except Exception as e: 
        return True,"系統無法分析此樣板，出現意外錯誤"

    new_file=f"media/salary_files/{full_name}_{employee_id}_{title}.xlsx"
    workbook.save(new_file)
    return False,new_file
