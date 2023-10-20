from datetime import date, timedelta
from Backend.models import Employee,Travel_Application,Clock_Correction_Application, Work_Overtime_Application, Salary,SalaryDetail,Leave_Param,Leave_Application, Clock,Project_Confirmation,Project_Job_Assign,Project_Employee_Assign
from urllib.parse import parse_qs
from django.forms.models import model_to_dict
from django.conf import settings
from django.http import HttpResponse
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from urllib.parse import quote
import os
from django.db.models import Sum
import math
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill




def employeeAssignFile(employee_assign_obj):

    client = employee_assign_obj.project_job_assign.project_confirmation.quotation.client or ""
    quotation_id = employee_assign_obj.project_job_assign.project_confirmation.quotation.quotation_id or ""
    construction_date = employee_assign_obj.construction_date or ""
    requisition = employee_assign_obj.project_job_assign.project_confirmation.quotation.requisition.client_name or ""
    completion_date =  employee_assign_obj.completion_date  or ""
    remark =  employee_assign_obj.remark or ""
    
    project_name = employee_assign_obj.project_job_assign.project_confirmation.quotation.project_name or ""
    location = employee_assign_obj.project_job_assign.location or ""  
    manuscript_return_date = employee_assign_obj.manuscript_return_date or ""

    vehicle_str = ' '.join(item.vehicle_id for item in employee_assign_obj.project_job_assign.vehicle.all())
    work_employee_str = ' '.join(item.full_name for item in employee_assign_obj.project_job_assign.work_employee.all())

    uploaded_files = employee_assign_obj.uploaded_files.all() 
    test_items_ary = employee_assign_obj.test_items_ary() 
    img_dict={"帶班主管簽名":"帶班主管","填表人簽名":"填表人"
              ,"業主的簽名":"業主","總經理簽名":"總經理","經理簽名":"經理",
              "清點人":"清點人","會點人":"會點人","交接人_簽名":"交接人",
              "被交接人(1)_簽名":"被交接人1","被交接人(2)_簽名":"被交接人2"
              }

    carry_eq =  employee_assign_obj.carry_equipments_ary()

    file_path = r'media/system_files/employee_assign_template.xlsx'
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    print("xx")

    try:
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            print(i)
            if i >=55:
                break
            for index, cell_value in enumerate(row, start=1):
                print(cell_value)
                if cell_value == "CLIENT":
                    sheet.cell(row=i, column=index, value=str(client))
                elif cell_value == "QUOTATION_ID":
                    sheet.cell(row=i, column=index, value=quotation_id)
                elif cell_value == "VEHICLE":
                    sheet.cell(row=i, column=index, value=vehicle_str)
                elif cell_value == "CONSTRUCTION_DATE":
                    sheet.cell(row=i, column=index, value=construction_date)
                elif cell_value == "REQUISITION":
                    sheet.cell(row=i, column=index, value=str(requisition))
                elif cell_value == "WORK_EMPLOYEE":
                    sheet.cell(row=i, column=index, value=work_employee_str)
                elif cell_value == "COMPLETION_DATE":
                    sheet.cell(row=i, column=index, value=completion_date)
                elif cell_value == "PROJECT_NAME":
                    sheet.cell(row=i, column=index, value=project_name)
                elif cell_value == "LOCATION":
                    sheet.cell(row=i, column=index, value=location)
                elif cell_value == "MANUSCRIPT_RETURN_DATE":
                    sheet.cell(row=i, column=index, value=manuscript_return_date)
                elif cell_value == "remark":
                    remark = "※現場未復原之設備: " + str(remark)
                    sheet.cell(row=i, column=index, value=remark)
                elif img_dict.get(cell_value) is not None:
                    #假如把文字替換
                    if not "交接人"  in cell_value :
                        sheet.cell(row=i, column=index, value="")

                    cell_value_name = img_dict.get(cell_value)
                    uploaded_file = uploaded_files.filter(name=cell_value_name).first()
                    if uploaded_file:
                        img_path= uploaded_file.file
                        file_path = os.path.join(settings.MEDIA_ROOT, str(img_path))
                        if os.path.exists(file_path):
                            with open(file_path, 'r') as f:
                                img = Image(file_path)
                                if  cell_value== "清點人" or cell_value=="會點人":
                                    img.width =79
                                    img.height =46
                                else:
                                    img.width =157
                                    img.height =92

                                sheet.add_image(img, f'{sheet.cell(row=i, column=index).coordinate}')
                        else:

                            print(f"找不到檔案: {file_path}")


                elif cell_value == "eq":
                    add_col = 0
                    next_index_row = 0
                    if  len(carry_eq)==0:
                        sheet.cell(row=i, column=index, value="")

                    for next_index, item in enumerate(carry_eq): #艾利克
                        # print(next_index_row)
                        # print("---")
                        # print(next_index)
                        # print(i+next_index_row+add_col,index+add_col)
                        sheet.cell(row=i+next_index_row, column=index+add_col, value= item["status"])
                        eq_name = item["equipment_id"]+" "+ item["equipment_name"]
                        sheet.cell(row=i+next_index_row, column=index+1+add_col, value=eq_name)
                        next_index_row+=1
                        if next_index==18:
                            add_col =12
                            next_index_row=0
                            # print("next +2")
                            # print(index+1+add_col,next_index_row)
                        if  next_index==37:
                            # print("xxx end")
                            break

                elif cell_value == "TEST_ITEM":
                    for next_index, item in enumerate(test_items_ary):
                        if next_index ==10:
                            break
                        sheet.cell(row=i+next_index, column=index, value= item["test_date"])
                        sheet.cell(row=i+next_index, column=index+3, value= item["test_location"])
                        sheet.cell(row=i+next_index, column=index+6, value= item["test_items"])
                        sheet.cell(row=i+next_index, column=12, value= item["format_and_voltage"])
                        sheet.cell(row=i+next_index, column=20, value= item["level"])
                        

                        
                #     sheet.cell(row=i, column=index, value=sum)
                # elif cell_value == "hide_key":
                #     if see:
                #         sheet.cell(row=i, column=index, value="客戶名稱").fill = grey_fill
                #     else:
                #         sheet.cell(row=i, column=index, value="")
                # elif cell_value == "hide_name":
                #     if five:
                #         sheet.cell(row=i, column=index, value=hide_client)
                #     else:

                #         if see:
                #             sheet.cell(row=i, column=index, value=hide_client).fill = grey_fill
                #         else:
                #             sheet.cell(row=i, column=index, value="")
                # elif cell_value == "sum_five":
                #     if five:                           
                #         five_sum = math.ceil(sum *0.05)
                #     else: 
                #         five_sum=0
                #     sheet.cell(row=i, column=index, value=five_sum)
                # elif cell_value == "all_sum":
                #     all_sum = sum + five_sum
                #     sheet.cell(row=i, column=index, value=all_sum)

    except ValueError as e:
        return True,"遇到欄位合併的錯誤"
    except Exception as e: 
        print(e)
        return True,"系統無法分析此樣板，出現意外錯誤"

    filename = f"【派工單】 {employee_assign_obj}-{requisition}-{project_name}.xlsx"
    quoted_filename = quote(filename)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] =  f'attachment; filename="{quoted_filename}"'
    workbook.save(response)
    return response
