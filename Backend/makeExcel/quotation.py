from datetime import date, timedelta
from Backend.models import Employee,Travel_Application,Clock_Correction_Application, Work_Overtime_Application, Salary,SalaryDetail,Leave_Param,Leave_Application, Clock,Project_Confirmation,Project_Job_Assign,Project_Employee_Assign
from urllib.parse import parse_qs
from django.forms.models import model_to_dict
from django.conf import settings
from django.http import HttpResponse
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from urllib.parse import quote
import os
from django.db.models import Sum
import math
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill




def quotationFile(quotation_obj,see,five):
    quotation_id = "" if quotation_obj.quotation_id is None else quotation_obj.quotation_id
    hide_client = "" if quotation_obj.client.client_name is None else quotation_obj.client.client_name
    requisition = "" if quotation_obj.requisition.client_name is None else quotation_obj.requisition.client_name
    project_name = "" if quotation_obj.project_name is None else quotation_obj.project_name
    tax_id =  quotation_obj.requisition.tax_id  or ""
    contact_person = quotation_obj.requisition.contact_person or ""

    get_business_assistant = quotation_obj.business_assistant_user 
    business_assistant_user_name=""
    business_assistant_user_mail=""
    business_tel =""
    if get_business_assistant :
        business_assistant_user_name,business_assistant_user_mail, business_tel=  get_business_assistant.info()

    address = quotation_obj.requisition.address or ""  
    tel = quotation_obj.requisition.tel or ""
    mobile = quotation_obj.requisition.mobile or ""
    fax = quotation_obj.requisition.fax or ""
    email = quotation_obj.requisition.email or ""
    quote_validity_period = quotation_obj.quote_validity_period or ""
    quote_date = quotation_obj.quote_date or ""
    

    if quote_validity_period:
        quote_validity_period=str(quote_validity_period) + "天"

    item_list = quotation_obj.Quotation_Work_Item_Number.all()

    if five:
        file_path = r'media/system_files/quotation_template_other.xlsx' # 維景 還沒放
    else:
        file_path = r'media/system_files/quotation_template.xlsx' # 艾力克電機

    try:
        workbook = load_workbook(filename=file_path)
    except Exception as e:
        return True,"檔案位置發生問題"


    sheet = workbook.active
    # chinese_format = NamedStyle(name='chinese_format', number_format="[$NT$-2]#,##0;[RED]-[$NT$-2]#,##0")
    # workbook.add_named_style(chinese_format)

    #放圖
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")


    if see:
        uploaded_files = quotation_obj.uploaded_files.all() 

        for i, file in enumerate(uploaded_files):
            file_path = os.path.join(settings.MEDIA_ROOT, str(file.file))
            if os.path.exists(file_path):
                with open(file_path, 'r') as f:
                    img = Image(file_path)
                    sheet.add_image(img, f'O{4+i}')
            else:
                sheet[f'O{4+i}'].value=f"伺服器找不到{file.name}、路徑{file.file}檔案"
                print(f"找不到檔案: {file_path}")

        #顯示對內對話紀錄
        internal_content ="" if quotation_obj.internal_content is None else quotation_obj.internal_content
        sheet['M2'].value = internal_content


    sum = 0
    five_sum = 0
    try:
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            
            for index, cell_value in enumerate(row, start=1):
                if cell_value == "CLIENT":
                    sheet.cell(row=i, column=index, value=str(requisition))
                elif cell_value == "QUOTATION_ID":
                    sheet.cell(row=i, column=index, value=quotation_id)
                elif cell_value == "quote_date":
                    sheet.cell(row=i, column=index, value=quote_date)
                elif cell_value == "TAX_ID":
                    sheet.cell(row=i, column=index, value=tax_id)
                elif cell_value == "MOBILE":
                    sheet.cell(row=i, column=index, value=mobile)
                elif cell_value == "CONTACT_PERSON":
                    sheet.cell(row=i, column=index, value=contact_person)
                elif cell_value == "BUSINESS_ASSISTANT":
                    sheet.cell(row=i, column=index, value=business_assistant_user_name)
                elif cell_value == "BUSINESS_ASSISTANT_mail":
                    sheet.cell(row=i, column=index, value=business_assistant_user_mail)
                elif cell_value == "PROJECT_NAME":
                    sheet.cell(row=i, column=index, value=project_name)
                elif cell_value == "TEL":
                    sheet.cell(row=i, column=index, value=tel)
                elif cell_value == "FAX":
                    sheet.cell(row=i, column=index, value=fax)
                elif cell_value == "EMAIL":
                    sheet.cell(row=i, column=index, value=email)
                elif cell_value == "ADDRESS":
                    sheet.cell(row=i, column=index, value=address)
                elif cell_value == "QUOTE_VALIDITY_PERIOD":
                    sheet.cell(row=i, column=index, value=quote_validity_period)
                elif cell_value == "BUSINESS_TEL":
                    sheet.cell(row=i, column=index, value=business_tel)
                elif cell_value == "ITEM_LIST":
                    if  len(item_list)==0:
                        sheet.cell(row=i, column=index, value="")
                        sheet.cell(row=i, column=index+1, value="")
                        sheet.cell(row=i, column=index+5, value="")
                        sheet.cell(row=i, column=index+6, value="")
                        sheet.cell(row=i, column=index+7, value="")
                        sheet.cell(row=i, column=index+8, value="")
                    else:
                        for next_index, item in enumerate(item_list): #艾利克
                            sheet.cell(row=i+next_index, column=index, value= next_index+1)
                            sheet.cell(row=i+next_index, column=index+1, value=item.work_item.item_name)
                            sheet.cell(row=i+next_index, column=index+5, value=item.work_item.contract_id)
                            sheet.cell(row=i+next_index, column=index+6, value=item.work_item.unit)
                            num =item.number                            
                            money=item.work_item.last_money_year()[0]
                            item_total_moeny=0
                            if str(money).isnumeric():
                                item_total_moeny = num*money
                            sheet.cell(row=i+next_index, column=index+7, value=num)
                            sheet.cell(row=i+next_index, column=index+8, value=money )
                            sheet.cell(row=i+next_index, column=index+9, value=item_total_moeny )
                            sum+=item_total_moeny
                elif cell_value == "SUM":
                    sheet.cell(row=i, column=index, value=sum)
                elif cell_value == "hide_key":
                    if see:
                        sheet.cell(row=i, column=index, value="客戶名稱").fill = grey_fill
                    else:
                        sheet.cell(row=i, column=index, value="")
                elif cell_value == "hide_name":
                    if five:
                        sheet.cell(row=i, column=index, value=hide_client)
                    else:

                        if see:
                            sheet.cell(row=i, column=index, value=hide_client).fill = grey_fill
                        else:
                            sheet.cell(row=i, column=index, value="")
                elif cell_value == "sum_five":
                    if five:                           
                        five_sum = math.ceil(sum *0.05)
                    else: 
                        five_sum=0
                    sheet.cell(row=i, column=index, value=five_sum)
                elif cell_value == "all_sum":
                    all_sum = sum + five_sum
                    sheet.cell(row=i, column=index, value=all_sum)

    except ValueError as e:
        return True,"遇到欄位合併的錯誤"
    except Exception as e: 
        print(e)
        return True,"系統無法分析此樣板，出現意外錯誤"

    filename = f"【報價單】 {quotation_id}-{requisition}-{project_name}.xlsx"
    quoted_filename = quote(filename)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] =  f'attachment; filename="{quoted_filename}"'
    workbook.save(response)
    return response
