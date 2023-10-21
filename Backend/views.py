import math
from django.shortcuts import render
from django.http import JsonResponse,HttpResponseNotAllowed,HttpResponseRedirect,HttpResponse,Http404
import json,datetime
from django.contrib.auth import update_session_auth_hash
import base64
from django.core.files.base import ContentFile
from django.core.exceptions import ObjectDoesNotExist
from Backend.models import LOCATION_CHOICES
from Backend.models import  AnnualLeave,ReferenceTable, Vehicle,Work_Item_Number,Travel_Application, ExtraWorkDay,Clock_Correction_Application, Work_Overtime_Application, Salary,SalaryDetail, Client,Leave_Application,Leave_Param,SysMessage,Approval_Target, Equipment, UploadedFile,Department,Quotation,ApprovalLog,Work_Item,ApprovalModel, Project_Confirmation, Employee, Project_Job_Assign,News,Clock,Project_Employee_Assign
from Backend.forms import ReferenceTableForm,VehicleForm,Travel_ApplicationForm,ExtraWorkDayForm, ClientForm, ClockCorrectionApplicationForm, WorkOvertimeApplicationForm, LeaveParamModelForm,LeaveApplicationForm,ProjectConfirmationForm,EquipmentForm,QuotationForm,DepartmentForm,Work_ItemForm,  EmployeeForm, ProjectJobAssignForm,NewsForm,Project_Employee_AssignForm
from django.contrib.auth.models import User,Group
from django.contrib.auth.forms import PasswordChangeForm
from urllib.parse import parse_qs
from urllib.parse import quote
from django.shortcuts import get_object_or_404
from django.forms.models import model_to_dict
from django.views import View
from .utils import Check_Permissions,convent_dict,convent_employee,convent_excel_dict,get_model_by_name
from Backend.makeExcel.salary import salaryFile
from Backend.makeExcel.quotation import quotationFile
from Backend.makeExcel.employee_assign import employeeAssignFile
from .salary_utils import create_salary
import openpyxl
from openpyxl import load_workbook

from openpyxl.utils import get_column_letter
from django.db.utils import IntegrityError
import random,os,re
from  django.conf import settings

from django.contrib.auth.mixins import UserPassesTestMixin

from datetime import time

from django.utils.text import get_valid_filename # 確保file檔名是合法的，不接受的符號會轉成可接受符號
import zipfile


class EmployeeAssignileView(View):

    def get(self, request, *args, **kwargs):
        print("進到get")
        obj_id = self.kwargs.get('id')
        
        if id:
            error_msg=""
            try:
                employee_assign_obj = Project_Employee_Assign.objects.get(id=obj_id)
                print("employee_assign_obj ",employee_assign_obj)
            except Http404:
                return JsonResponse({"error": "找不到相應的ID obj"}, status=400)
            except Exception as e:
                print(e)
                return JsonResponse({"error": str(e)}, status=400)
            
            return employeeAssignFile(employee_assign_obj)

        else:
            print("error 找不到相應的ID obj")
            return JsonResponse({"error": "找不到相應的ID obj"}, status=400)

class Work_ItemFileView(View):

    def get(self, request, *args, **kwargs):
        file_path = r'media/file_form_template/工項管理樣板.xlsx' 
        get_obj = Work_Item.objects.all()
        workbook = load_workbook(file_path)
        sheet = workbook.active
        last_row = sheet.max_row
        for i, item in enumerate(get_obj, start=last_row+1):
            price, year = item.last_money_year()
            sheet.cell(row=i, column=1, value=item.item_id)
            sheet.cell(row=i, column=2, value=item.item_name)
            sheet.cell(row=i, column=3, value=item.contract_id)
            sheet.cell(row=i, column=4, value=item.requisition.name)
            sheet.cell(row=i, column=5, value=item.unit)
            sheet.cell(row=i, column=6, value=year)
            sheet.cell(row=i, column=7, value=price)        

        new_path =r"media/file_form_template/工項管理匯出.xlsx"
        workbook.save(new_path)
        with open(new_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={os.path.basename(new_path)}'
            return response


class QuotationFileView(View):

    def get(self, request, *args, **kwargs):
        obj_id = self.kwargs.get('id')
        see = self.kwargs.get('see')
        five = self.kwargs.get('five')
        
        error_msg=""
        try:
            quotaion_obj = Quotation.objects.get(id=obj_id)
            print("quotaion_obj ",quotaion_obj)
        except Http404:
            return JsonResponse({"error": "找不到相應的ID obj"}, status=400)
        except Exception as e:
            print(e)
            return JsonResponse({"error": str(e)}, status=400)
        
        return quotationFile(quotaion_obj,see,five)

       
        
class SalaryFileView(View):

    def get(self, request, *args, **kwargs):
        year, month, user = self.kwargs.get('year'), self.kwargs.get('month'), self.kwargs.get('user')
        use_type =  self.kwargs.get('use_type')
        use_type = bool(int(use_type))
        title=""
        if use_type:
            title = "薪資條"
        else:
            title = "激勵性獎金"
        
        if user==0:#全員
            getEm = Employee.objects.filter(user__is_active=True)
            error_msg=""
            file_paths=[]
            for employee in getEm:
                try:
                    get_salary = Salary.objects.get(user=employee, year=year, month=month)
                except Http404:
                    error_msg +=f"找不到{employee.full_name}的當月薪資單 \n"
                    continue
                except Exception as e:
                    error_msg +=f"找不到{employee.full_name} 其他錯誤，如沒有部門 \n"
                    continue
                       
                error,obj = salaryFile(get_salary,use_type)

                if error:
                    error_msg += f"{obj} \n"
                else:      
                    file_paths.append({"file_name":obj,"name":f"{year}_{month}_{title}/{employee.full_name}_{year}_{month}_{title}.xlsx"})
                
            filename = f'media/salary_files/全員_{year}_{month}_{title}.zip'
            quoted_filename = quote(filename)

            with zipfile.ZipFile(filename, 'w') as archive:
                for file_info in file_paths:
                    file_name = file_info["file_name"]
                    name = file_info["name"]
                    archive.write(file_name, arcname=name)

            with open(filename, 'rb') as file:
                response = HttpResponse(file, content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="{quoted_filename}"'
                response['Content-Length'] = os.path.getsize(filename)

            print(error_msg)
            return response

        else:
            try:
                get_salary = Salary.objects.get(user=user, year=year, month=month)
            except Http404:
                return JsonResponse({"error": "找不到相應的ID obj"}, status=400)
       
            error,obj = salaryFile(get_salary,use_type)

            if error:
                return JsonResponse({"error":obj},status=400)
         

            with open(obj, 'rb') as file:
                filename = f"{get_salary.user.full_name}_{year}_{month}_{title}.xlsx"
                print(filename)
                response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={quote(filename)}'

            return response



class SalaryDetailView(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"薪水管理")

    def delete(self, request, *args, **kwargs):
        dict_data = convent_dict(request.body)

        id = int(dict_data.get('itemid'))
        try:
            get_obj = get_object_or_404(SalaryDetail, id=id)
        except Http404:
            return JsonResponse({"error": "找不到相應的明細 obj"}, status=400)
        get_obj.delete()
        return JsonResponse({"ok":"ok"},status=200)

    def put(self, request, *args, **kwargs):
        dict_data = convent_dict(request.body)

        id = int(dict_data.get('itemid'))
        name = dict_data.get('name') 
        deduction = dict_data.get('deduction')
        tax_deduction = dict_data.get('tax_deduction')
        five = dict_data.get('five')
        adjustmentAmount = int(dict_data.get('adjustmentAmount'))

        if adjustmentAmount <0:
            return JsonResponse({"error": "調整金額需大於0，如需扣款請勾選為扣款項"}, status=400)

        try:
            get_obj = get_object_or_404(SalaryDetail, id=id)
        except Http404:
            return JsonResponse({"error": "找不到相應的ID obj"}, status=400)

        get_obj.set_name_and_adjustment_amount(name, adjustmentAmount,deduction,tax_deduction=tax_deduction,five=five)  

        return JsonResponse({"ok":"ok"},status=200)

    def post(self, request, *args, **kwargs):
        dict_data = convent_dict(request.body)
        moneyValue = dict_data.get('moneyValue',0)
        if int(moneyValue) <= 0:
            return JsonResponse({"error": "金額需大於0"}, status=400)

        name = dict_data.get('name')
        deductionValue = dict_data.get('deduction')
        tax_deduction = dict_data.get('tax_deduction')
        five = dict_data.get('five')
        year, month, user = self.kwargs.get('year'), self.kwargs.get('month'), self.kwargs.get('user')
        try:
            get_salary =Salary.objects.get(user=user, year=year, month=month)
        except Http404:
            return JsonResponse({"error": "找不到相應的ID obj"}, status=400)

        SalaryDetail.objects.create(
            salary=get_salary,
            name=name,
            system_amount=moneyValue,  
            adjustment_amount=moneyValue,
            deduction=deductionValue,
            tax_deduction=tax_deduction,
            five=five,
        )
        return JsonResponse({"ok":"ok"},status=200)


        

class SalaryListView(UserPassesTestMixin,View):
    def test_func(self):        
        return Check_Permissions(self.request.user,"薪水管理")

    def post(self, request, *args, **kwargs):
        year = kwargs.get('year')
        month = kwargs.get('month')
        user_id = kwargs.get('user_id')

        errors=""

        if user_id:
            employee = Employee.objects.get(id=user_id)
            employee_location = employee.location_city
            if employee_location =="":
                errors+=f"{employee.full_name} 未輸入系統計算城市欄位資料，請通知管理部調整。<br>"
            else:
                try:
                    create_salary(employee, year, month)
                except Exception as e:
                    errors+=f"{employee.full_name} 薪資處理有問題 {e}<br>"

        else:
            employees = Employee.objects.filter(user__is_active=True).exclude(user__username='admin')
            for employee in employees:
                employee_location = employee.location_city
                if employee_location =="":
                    errors+=f"{employee.full_name} 未輸入系統計算城市欄位資料，請通知管理部調整。<br>"
                else:
                    try:
                        create_salary(employee, year, month)
                    except Exception as e:
                        errors+=f"{employee.full_name} 薪資處理有問題 {e}<br>"

            #刪除淘汰的員工
            del_employees = Employee.objects.filter(user__is_active=False)
            salaries_to_delete = Salary.objects.filter(user__in=del_employees)
            salaries_to_delete.delete()
        if errors:
            return JsonResponse({"error":errors+"<br>其餘員工計算完成。"},status=400)
        else:
            return JsonResponse({"ok":"ok"},status=200)

class SysMessage_API(View):
    def post(self, request, *args, **kwargs):
        data = request.POST
        id = data.get('id')  
        getobj = SysMessage.objects.get(id=id)
        if request.user.employee != getobj.Target_user:
            return HttpResponse("User does not have permission to delete this message", status=403)
         
        getobj.delete()
        return HttpResponse(200)


class Approval_View_Process(View):
    def post(self,request):
        new_Approval=None
        
        try:
            id = request.POST.get('id')
            modeltext = request.POST.get('model')

            getmodel = get_model_by_name(modeltext)
            if getmodel is  None:
                return JsonResponse({"error":"找不到model"},status=400)

            try:
                get_obj = get_object_or_404(getmodel, id=id)
            except Http404:
                return JsonResponse({"error": "找不到相應的ID obj"}, status=400)
            
            model_name = {# 對應Approval_Target的name
                'project_confirmation': 'Project_Confirmation',
                'job_assign': 'Project_Job_Assign',
                'Project_Employee_Assign': 'Project_Employee_Assign',
                'Work_Overtime_Application': 'Work_Overtime_Application',
                'Leave_Application': 'Leave_Application',
                'Clock_Correction_Application': 'Clock_Correction_Application',
                'Travel_Application': 'Travel_Application',
            }

            try:
                get_Approval_Target = get_object_or_404(Approval_Target, name=model_name.get(modeltext))
            except Http404:
                return JsonResponse({"error": "找不到相應的簽核目標"}, status=400)
            
            get_order = get_Approval_Target.approval_order

            if modeltext =="Leave_Application": #請假單
                get_user = get_obj.substitute
                if get_user :
                    get_user_id= get_user.id
                    get_order.insert(0, get_user_id ) #先給代理人簽核
                else:
                    return JsonResponse({"error": "請選擇代理人"}, status=400)

            new_Approval= ApprovalModel.objects.create(target_approval=get_Approval_Target,order=get_order)
            if get_obj.Approval !=None:
                get_obj.Approval.delete()
            get_obj.Approval=new_Approval
            get_obj.save()

            employee_id = new_Approval.target_approval.approval_order[0]
            if  employee_id !="x":
                sender = Employee.objects.get(id=employee_id)
                SysMessage.objects.create(Target_user=sender,content=f"您有一筆 {new_Approval.target_approval.get_name_display()} 單需要簽核")
            else:
                print("new_Approval")
                print(new_Approval.get_created_by)

                senders = new_Approval.get_created_by.departments.employees.filter(user__groups__name='主管')
                for sender in senders:
                    SysMessage.objects.create(Target_user=sender,content=f"您有一筆 {new_Approval.target_approval.get_name_display()} 單需要簽核")


            return JsonResponse({"success":"成功"},status=200)
        except Exception as e:
            if new_Approval !=None: #刪除出錯的簽核執行obj
                new_Approval.delete()
            print(str(e))
            return JsonResponse({"error": f"系統發生錯誤, {str(e)} "}, status=500)

    def delete(self,request):
        data = request.body.decode('utf-8')
        parsed_data = parse_qs(data)
        id = parsed_data.get('id', [None])[0]
        modeltext = parsed_data.get('model', [None])[0]
        getmodel = get_model_by_name(modeltext)

        if getmodel is  None:
            return JsonResponse({"error":"找不到model"},status=400)

        try:
            get_obj = get_object_or_404(getmodel, id=id)
        except Http404:
            return JsonResponse({"error": "找不到相應的ID obj"}, status=400)
        
        user = request.user.employee
        approval_obj = get_obj.Approval

        if approval_obj.current_status == 'in_process':
            approval_obj.send_message_to_related_users(f"{get_obj.get_show_id()} 單被{user.full_name}收回簽核")
            approval_obj.delete()

            return JsonResponse({"message": "删除成功"}, status=200)
        elif approval_obj.current_status == 'completed':
            if request.user.groups.filter(name='主管').exists():
                approval_obj.send_message_to_related_users(f"{get_obj.get_show_id()} 單被{user.full_name}收回簽核")
                approval_obj.delete()
                # get_obj.Approval=None
                # get_obj.save()
                return JsonResponse({"message": "删除成功"}, status=200)
            else:
                return JsonResponse({"error": "只有主管才可處理"}, status=403)
        else:
            return JsonResponse({"error": "操作不允許"}, status=403)
 


class Approval_Process_Log(View):
    def post(self,request):
        status = request.POST.get('status')
        feedback = request.POST.get('feedback')
        approval_id = request.POST.get('Approval_id')

        if status not in ["approved", "rejected"]:
            return JsonResponse({"error":"請選擇簽核狀態"},status=400)
        
        if approval_id ==None: 
            return JsonResponse({"error":"請選擇approval"},status=400)
        
        #尋找實體
        try:
            approval_instance = get_object_or_404(ApprovalModel, id=approval_id)
        except Http404:
            return JsonResponse({"error": "找不到相應的ApprovalModel"}, status=400)

        ApprovalLog.objects.create(
            approval=approval_instance,
            user=request.user.employee,
            content=feedback,
        )
        approval_instance.update_department_status(status)
        
        if status:
            return JsonResponse({"data":"ok"},status=200)

class Department_View(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"管理部管理",self.request.method,"管理部查看") 

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = DepartmentForm(dict_data)
        if form.is_valid():

            get_object=Department.objects.get(id=dict_data['id'])
            if "parent_department" in dict_data:
                Equipment_nstance = Department.objects.get(pk=int(dict_data["parent_department"]))
                get_object.parent_department = Equipment_nstance
                del dict_data["parent_department"]
            
            get_object.update_fields_and_save(**dict_data)

            Department.objects.get(id=dict_data['id']).update_fields_and_save(**dict_data)
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)

    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            Department.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)


    def post(self,request):
        form = DepartmentForm(request.POST)

        if form.is_valid():
            newobj =form.save()
            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)


    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Department, id=id)
        data = model_to_dict(data)

        return JsonResponse({"data":data}, status=200,safe = False)


class Equipment_View(View):

    def put(self,request):
        dict_data = convent_dict(request.body)

        form = EquipmentForm(dict_data)
        if form.is_valid():
            Equipment.objects.get(id=dict_data['id']).update_fields_and_save(**dict_data)
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)

    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            Equipment.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)


    def post(self,request):
        form = EquipmentForm(request.POST)

        if form.is_valid():
            newobj =form.save()

            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)


    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Equipment, id=id)
        data = model_to_dict(data)
        data['abnormal_img'] = data['abnormal_img'].url if data['abnormal_img']   else ''
        data['buy_img'] = data['buy_img'].url if data['buy_img']  else ''
        print("dict_data")
        print(data)
        return JsonResponse({"data":data}, status=200,safe = False)

class ReferenceTable_View(UserPassesTestMixin,View):

    def test_func(self):
        return Check_Permissions(self.request.user,"管理部管理")

    def put(self,request):
        dict_data = convent_dict(request.body)

        form = ReferenceTableForm(dict_data)
        if form.is_valid():
            ReferenceTable.objects.get(id=dict_data['id']).update_fields_and_save(**dict_data)
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)

    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            ReferenceTable.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)


    def post(self,request):
        form = ReferenceTableForm(request.POST)
        if form.is_valid():
            location_city_residence = form.cleaned_data['location_city_residence']
            location_city_business_trip = form.cleaned_data['location_city_business_trip']
            name = form.cleaned_data['name']
            if ReferenceTable.objects.filter(
                        location_city_residence=location_city_residence,
                        location_city_business_trip=location_city_business_trip,
                        name=name
                    ).exists():
                return JsonResponse({"error": f"{name}參照表-居住地:{location_city_residence} 目的地:{location_city_business_trip} 已經有該筆紀錄"}, status=400)


            newobj =form.save()
            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)


    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(ReferenceTable, id=id)
        data = model_to_dict(data)        
        if data["amount"]=="":
            data["amount"] =0
        return JsonResponse({"data":data}, status=200,safe = False)

#派工單
class Project_Employee_Assign_View(UserPassesTestMixin,View):
    def test_func(self):        
        return Check_Permissions(self.request.user,"工程派工單管理",self.request.method,"工程派工單查看")  

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = Project_Employee_AssignForm(dict_data)
        if form.is_valid():
            get_Project_Employee_Assign =Project_Employee_Assign.objects.get(id=dict_data['id'])
            # print(dict_data)
            if "vehicle" in dict_data:
                related_project_job_assign = get_Project_Employee_Assign.project_job_assign
                vehicle_id = dict_data["vehicle"]
                vehicle_id_list = [int(item) for item in vehicle_id]
                related_project_job_assign.vehicle.set(vehicle_id_list)
                related_project_job_assign.save()
                del dict_data["vehicle"]

            if "project_job_assign" in dict_data:
                project_job_assign_instance = Project_Job_Assign.objects.get(pk=int(dict_data["project_job_assign"]))
                get_Project_Employee_Assign.project_job_assign = project_job_assign_instance
                del dict_data["project_job_assign"]

            get_Project_Employee_Assign.update_fields_and_save(**dict_data)

            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)


    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            obj = Project_Employee_Assign.objects.get(id=dict_data['id'])
            if obj.Approval:
                obj.Approval.delete()
            obj.delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)


    def post(self,request):
        print(request.POST)
        form = Project_Employee_AssignForm(request.POST)

        if form.is_valid():
            newobj =form.save()
            related_project_job_assign = newobj.project_job_assign
            vehicle = request.POST.getlist("vehicle")
       
            vehicle_id_list = [int(item) for item in vehicle]
            related_project_job_assign.vehicle.set(vehicle_id_list)
            related_project_job_assign.save()


            newobj.save()
            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)


    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Project_Employee_Assign, id=id)
        location = data.project_job_assign.location
        carry_equipments_str=data.carry_equipments_ary()
        vehicle_id_list = []
        for item in data.project_job_assign.vehicle.all():
            vehicle_id_list.append(item.id)

        uploaded_files = data.uploaded_files.all() 
        uploaded_files_dict_list = []
        for file in uploaded_files:
            file_dict = model_to_dict(file)
            file_dict["file"] = file.file.url
            uploaded_files_dict_list.append(file_dict)

        get_id=data.get_show_id()
        data = model_to_dict(data)
        data["carry_equipments_str"] =carry_equipments_str
        data["show_id"] = get_id
        data["uploaded_files"] = uploaded_files_dict_list
        data["location"] = location
        data["vehicle"] = vehicle_id_list
        data['last_excel'] = data['last_excel'].url if data['last_excel']  else None

        if  data['enterprise_signature']:
            data['enterprise_signature'] = data['enterprise_signature'].url
        else:
            data['enterprise_signature'] = None
        # print("dict_data")
        # print(data)

        return JsonResponse({"data":data}, status=200,safe = False)

class New_View(View):

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = NewsForm(dict_data)
        if form.is_valid():
            News.objects.get(id=dict_data['id']).update_fields_and_save(**dict_data)
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)


    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            News.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)


    def post(self,request):
        form = NewsForm(request.POST)

        if form.is_valid():
            newobj = form.save()
            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)


    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(News, id=id)
        data = model_to_dict(data)
        print("dict_data")
        print(data)
        data['attachment'] = data['attachment'].url if data['attachment']  else None

        return JsonResponse({"data":data}, status=200,safe = False)



class Quotation_View(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"業務部管理",self.request.method,"業務部查看") 
    def put(self,request):
        dict_data = convent_dict(request.body)
        form = QuotationForm(dict_data)

        if form.is_valid():
            getQuotation = Quotation.objects.get(id=dict_data['id'])

            work_item_objs = getQuotation.Quotation_Work_Item_Number.all()

            Work_Item_Number_ids = dict_data.get('Work_Item_Number_id', [])


            for obj in work_item_objs:
                if obj.id not in Work_Item_Number_ids:
                    obj.delete()

            if "work_item_id" in dict_data:
                # print(work_item_ids,work_item_numbers,Work_Item_Number_ids)
                work_item_ids = dict_data['work_item_id'] 
                work_item_numbers = dict_data['work_item_number'] 
                for i in range(len(Work_Item_Number_ids)):
                    work_item_id = work_item_ids[i]
                    work_item_number = work_item_numbers[i]
                    Work_Item_Number_id = Work_Item_Number_ids[i]
                    matching_obj = Work_Item_Number.objects.filter(id=Work_Item_Number_id).first()
                    try:
                        work_obj = get_object_or_404(Work_Item, id=work_item_id)                    
                    except Http404:
                        return JsonResponse({"error":"找不到工項，請先重新整理看看。"},status=400)
                    if matching_obj:
                        matching_obj.number = work_item_number
                        matching_obj.work_item = work_obj
                        matching_obj.save()
                    else:
                        Work_Item_Number.objects.create(
                            quotation=getQuotation,
                            work_item=work_obj,
                            number=work_item_number
                        )                

            if "client" in dict_data:
                get_obj = Client.objects.get(pk=int(dict_data["client"]))
                getQuotation.client = get_obj
                del dict_data["client"]

            if "business_assistant_user" in dict_data:
                get_obj = Employee.objects.get(pk=int(dict_data["business_assistant_user"]))
                getQuotation.business_assistant_user = get_obj
                del dict_data["business_assistant_user"]

            if "requisition" in dict_data:
                get_obj = Client.objects.get(pk=int(dict_data["requisition"]))
                getQuotation.requisition = get_obj
                del dict_data["requisition"]

            getQuotation.update_fields_and_save(**dict_data)
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)


    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            quotation_obj = Quotation.objects.get(id=dict_data['id']) 
            if hasattr(quotation_obj,"work_item") and quotation_obj.work_item is not None:
                print("存在")
                print("quotation_obj.work_item: ", quotation_obj.work_item.all())
                for i in quotation_obj.work_item.all():
                    for n in i:
                        Work_Item_Number.objects.get(id=n.Work_Item_Number.id).delete()

            quotation_obj.delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)


    def post(self,request):
        form = QuotationForm(request.POST)
        
        work_item_id = request.POST.getlist('work_item_id')
        work_item_number = request.POST.getlist('work_item_number')
        if form.is_valid():
            newobj = form.save()
            for i, n in zip(work_item_id, work_item_number):
                work_item = Work_Item.objects.get(id=i)
                Work_Item_Number.objects.create(
                    quotation = newobj,
                    work_item = work_item,
                    number = n
                )
            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)


    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Quotation, id=id)
        get_id = data.get_show_id()

        uploaded_files = data.uploaded_files.all() 
        uploaded_files_dict_list = []
        for file in uploaded_files:
            file_dict = model_to_dict(file)
            file_dict["file"] = file.file.url
            uploaded_files_dict_list.append(file_dict)
            

        work_item_list = []
        for item in data.Quotation_Work_Item_Number.all():
            work_item_list.append({"id":item.id,"number":item.number,"work_item_id":item.work_item.id})
            
        client_name=data.client.client_name if data.client else None
        requisition_name=data.requisition.client_name if data.requisition else None
        data = model_to_dict(data)
        print("dict_data")
        print(data)
        data['last_excel'] = data['last_excel'].url if data['last_excel']  else None
        data["uploaded_files"]=uploaded_files_dict_list
        data['work_item'] = work_item_list
        data['quotation_id'] = get_id
        data['client_name'] = client_name
        data['requisition_name'] = requisition_name

        return JsonResponse({"data":data}, status=200,safe = False)


class Work_Item_View(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"業務部管理",self.request.method,"業務部查看") 

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = Work_ItemForm(dict_data)
        if form.is_valid():
            getWork_Item =Work_Item.objects.get(id=dict_data['id'])
            if "requisition" in dict_data:
                get_obj = Client.objects.get(pk=int(dict_data["requisition"]))
                getWork_Item.requisition = get_obj
                del dict_data["requisition"]

            getWork_Item.update_fields_and_save(**dict_data)
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)


    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            Work_Item.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)


    def post(self,request):
        form = Work_ItemForm(request.POST)

        if form.is_valid():
            newobj = form.save()
            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)


    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Work_Item, id=id)
        #找尋使用過的報價單
        work_item_number = Work_Item_Number.objects.filter(work_item=data)

        quotations_info=[]
        for wn in work_item_number:
            user_name=""
            if wn.quotation.business_assistant_user:
                user_name = wn.quotation.business_assistant_user.full_name
            quotation_info = {
                "quotation_id": wn.quotation.quotation_id,
                "requisition": wn.quotation.requisition.client_name,
                "project_name": wn.quotation.project_name,
                "quote_date": wn.quotation.quote_date,
                "business_assistant": user_name,
                "number": wn.number,
            }
            quotations_info.append(quotation_info)



        data = model_to_dict(data)

        data['quotations_info'] = quotations_info
        print("dict_data")
        print(data)

        return JsonResponse({"data":data}, status=200,safe = False)

class Client_View(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"業務部管理",self.request.method,"業務部查看") 
    
    def put(self,request):
        dict_data = convent_dict(request.body)
        print(dict_data)
        form = ClientForm(dict_data)
        if form.is_valid():
            Client.objects.get(id=dict_data['id']).update_fields_and_save(**dict_data)
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)

    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            Client.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)

    def post(self,request):
        get_person = request.POST.getlist("person")
        # get_address = request.POST.getlist("address")

        form = ClientForm(request.POST)
        if form.is_valid():
            newobj = form.save()
            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)

    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Client, id=id)
        data = model_to_dict(data)
        return JsonResponse({"data":data}, status=200,safe = False)
    
class ExtraWorkDay_View(UserPassesTestMixin,View):

    def test_func(self):
        return Check_Permissions(self.request.user,"管理部管理")

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = ExtraWorkDayForm(dict_data)
        if form.is_valid():
            ExtraWorkDay.objects.get(id=dict_data['id']).update_fields_and_save(**dict_data)
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)

    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            ExtraWorkDay.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)

    def post(self,request):
        form = ExtraWorkDayForm(request.POST)
        if form.is_valid():
            newobj = form.save()
            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)

    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(ExtraWorkDay, id=id)
        data = model_to_dict(data)
        return JsonResponse({"data":data}, status=200,safe = False)
    



class IMGUploadView(View):
    def post(self, request):
        image_file = request.FILES.get('upload_img')
        if image_file !=None:
            random_number = str(random.randint(1, 99))
            new_file_name = f"{random_number}_{image_file.name}"
            file_path = rf'{settings.MEDIA_ROOT}/news_uploads/{new_file_name}'

            with open(file_path, 'wb') as f:
                for chunk in image_file.chunks():
                    f.write(chunk)

            return JsonResponse({"url":f"/media/news_uploads/{new_file_name}"},status=200)

        return HttpResponse(status=500)





#匯入檔案
class FileUploadView(View):
   def post(self, request, *args, **kwargs):
        from datetime import datetime
        modelstr = self.kwargs['model']
        uploaded_file = request.FILES.get('fileInput')

        print(modelstr)
        if uploaded_file ==None:
             return JsonResponse({'error': '請上傳檔案'}, status=400)
        #檢查type，回傳上傳正確檔案

        workbook = load_workbook(uploaded_file)
        worksheet = workbook.active
        get_dicts,get_model = convent_excel_dict(worksheet,modelstr)
        error_str=""
        print("get_dicts: " , get_dicts)
        for i,get_dict in enumerate(get_dicts):
            try:
                if get_model == Project_Confirmation:
                    print("get_dict: ",get_dict['quotation'])
                    id_object = Quotation.objects.get(quotation_id=get_dict['quotation'])
                    get_model.objects.create(
                        quotation = id_object,
                        order_id = get_dict['order_id'],    
                        c_a = get_dict['c_a'],    
                        turnover = get_dict['turnover'],    
                        remark = get_dict['remark']
                    )
                    print("id_object: ",id_object)
                elif get_model == Project_Job_Assign :
                    print("get_dict: ",get_dict['project_confirmation'])
                    id_str = get_dict['project_confirmation']
                    id_num = int(id_str[id_str.rindex("-")+1:])
                    print("id_num: ", id_num)
                    id_object = Project_Confirmation.objects.get(id=id_num)
                    get_model.objects.create(
                        project_confirmation = id_object,
                        attendance_date = get_dict['attendance_date'],    
                        location = get_dict['location'],    
                        remark = get_dict['remark']
                    )
                    print("工確單: ",id_object)
                elif get_model == Project_Employee_Assign:
                    print("get_dict: ",get_dict['project_job_assign'])
                    id_str = get_dict['project_job_assign']
                    id_num = int(id_str[id_str.rindex("-")+1:])
                    print("id_num: ", id_num)
                    id_object = Project_Job_Assign.objects.get(id=id_num)
                    get_model.objects.create(
                        project_job_assign = id_object,
                        construction_date = get_dict['construction_date'],    
                        completion_date = get_dict['completion_date'],    
                        manuscript_return_date = get_dict['manuscript_return_date'],
                        remark = get_dict['remark'],
                    )
                    print("派任計畫: ",id_object)
                elif get_model == Employee:
                    print("員工處理")
                    # print(get_dicts)
                    data = get_dict
                    username=data['employee_id']
                    id_number=data['id_number']
                    try:
                        user = User.objects.create_user(username=username, password=id_number)
                    except IntegrityError:
                        error_str += f"第{i+1}欄資料:ID {username}已存在過資料庫或是曾使用該帳號過。<br>"
                        continue

                    department_ary = data['departments'].split("/")
                    department=None 
                    if len(department_ary)==2:
                        try:
                            department = Department.objects.filter(belong_to_company=department_ary[0], department_name=department_ary[1]).first()
                        except Department.DoesNotExist:
                            error_str += f"第{i+1}欄資料:ID {username}建立，但部門無法搜尋，請在網頁操作。<br>"
                            pass
                    else:
                        error_str += f"第{i+1}欄資料:ID {username}建立，但部門格式錯誤，請在網頁操作。<br>"

                    location_city = data['location_city']
                    if  not location_city in dict(LOCATION_CHOICES):
                        error_str += f"第{i+1}欄資料:ID {username}建立，但找不到對應的居住城市(系統計算)，請在網頁操作。<br>"
                        location_city   =""

                    start_work_date = data['start_work_date']

                    if isinstance(start_work_date, str):
                        try:
                            start_work_date = datetime.strptime(start_work_date, '%Y/%m/%d').date()
                        except ValueError:
                            error_str += f"第{i+1}欄資料:ID {username}建立，無法解析到職日，請在網頁操作。<br>"
                            start_work_date = None

                    try:
                        employee = Employee(
                            full_name=data['full_name'],
                            employee_id=data['employee_id'],
                            id_number=data['id_number'],
                            gender=data['gender'],
                            blood_type=data['blood_type'],
                            departments=department,
                            position=data['position'],
                            start_work_date=start_work_date,
                            location_city=location_city,
                            user=user  
                        )
                        employee.save()
                        
                    except Exception as e:
                        user.delete()
                        error_str += f"第{i+1}欄資料:ID {username}建立<b>失敗</b>原因:{e}。<br>"
                elif get_model == Work_Item:
                    data = get_dict
                    # print(data)
                    existing_item = Work_Item.objects.filter(item_id=data['item_id']).first()
                    str_date = data['date'] or ""
                    
                    money = data['money'] or -1


                    requisition_obj = None
                    try:
                        requisition_obj = Client.objects.get(client_name=data['requisition'])
                    except Client.DoesNotExist:
                        error_str += f"第{i+1}欄資料:編號:{data['item_id']}建立，找不到客戶-{data['requisition']}後續請在網頁操作。<br>"

                    if str_date!="":
                        if  not isinstance(str_date, str):
                            str_date = str_date.strftime('%Y/%m/%d')
                        new_year_money = {'year':str_date, 'price': money} 
                    item_name = str(data['item_name'])

                    if existing_item:
                        existing_item.item_name =item_name
                        existing_item.requisition = requisition_obj
                        existing_item.contract_id =data['contract_id']
                        existing_item.unit = data['unit']
                        year_money_list = json.loads(existing_item.year_money)
                        add_data = True
                        if str_date=="":
                            existing_item.save()
                            continue                        
                        for entry in year_money_list:
                            # print(entry["year"],str_date,entry["price"],money)
                            if entry['year'] == str_date and entry['price'] == money:
                                add_data = False
                                break
                        if add_data:
                            year_money_list.append(new_year_money)
                            existing_item.year_money = json.dumps(year_money_list)
                        existing_item.save()
                    else:
                        if str_date!="":
                            new_year_moeny_str = json.dumps([new_year_money])
                        else:
                            new_year_moeny_str="[]"
                        new_item = Work_Item(
                            item_id=data['item_id'],
                            contract_id=data['contract_id'],
                            item_name=item_name,
                            requisition=requisition_obj,
                            unit=data['unit'],
                            year_money= new_year_moeny_str# 更新 year_money
                        )
                        new_item.save()


            
            except IntegrityError as e: #錯誤發生紀錄，傳給前端
                error_str+=f"第{i+1}欄資料錯誤<br>"
            except Quotation.DoesNotExist:
                error_str+=f"第{i+1}欄資料 找不到對應的找不到關聯報價單編號<br>"
            except Project_Confirmation.DoesNotExist:
                error_str+=f"第{i+1}欄資料 找不到對應的找不到工程確認單編號<br>"
            except Project_Job_Assign.DoesNotExist:
                error_str+=f"第{i+1}欄資料 找不到對應的找不到關聯報價單編號找不到關派任計畫編號<br>"
        if error_str=="":
            return JsonResponse({'message': '上傳成功'},status=200)
        else:
            return JsonResponse({'error': "<h4>異常欄位:</h4>\n"+error_str+"<br>其餘上傳成功，再次上傳請小心已經上傳過的資料。"}, status=400)




class FormUploadFileView(View):
   def post(self, request):
        getname = request.POST.get("name")
        getmodal = request.POST.get("modal")
        getid = request.POST.get("id")
        ManyToManyProcess = request.POST.get("ManyToManyProcess",False)
        filename = request.POST.get("file_name","")
        uploaded_file = request.FILES.get(getname)

        if getmodal =="Project_Employee_Assign":
            uploaded_file = request.POST.get("uploaded_files")
            base64_data = uploaded_file.split(',')[1]
            binary_data = base64.b64decode(base64_data)
            uploaded_file=ContentFile(binary_data, name=f'signature_{getid}_{filename}.png')


        #如果是多對多，就用另一種
        print("getmodel:", getmodal)
        print("getname:", getname)
        print("getid:", getid)
        print("uploaded_file:", uploaded_file)
        print("ManyToManyProcess:", ManyToManyProcess)
        print("filename:", filename)
        
        if(uploaded_file):
            match getmodal:
                case "project_confirmation":
                    model=Project_Confirmation.objects.get(id=getid)
                case "job_assign":
                    model=Project_Job_Assign.objects.get(id=getid)
                case "employee_assign":
                    model=Project_Employee_Assign.objects.get(id=getid)
                case "Project_Employee_Assign": #雖然有一樣的了，但仍獨立處理
                    model=Project_Employee_Assign.objects.get(id=getid)
                case "news":
                    model=News.objects.get(id=getid)
                case "Equipment":
                    model=Equipment.objects.get(id=getid)
                case "employee":
                    model=Employee.objects.get(id=getid)
                case "Quotation":
                    model=Quotation.objects.get(id=getid)
                case "Leave_Application":
                    model=Leave_Application.objects.get(id=getid)
                case "travel_application":
                    model=Travel_Application.objects.get(id=getid)
                case "Work_Overtime_Application":
                    model=Work_Overtime_Application.objects.get(id=getid)
                case "Clock_Correction_Application":
                    model=Clock_Correction_Application.objects.get(id=getid)
                case _:
                    return JsonResponse({"data":"no the modal"}, status=400,safe=False)

            if ManyToManyProcess:
                related_field = getattr(model, getname)
                uploaded_file_obj = UploadedFile.objects.create(name=filename, file=uploaded_file)
                related_field.add(uploaded_file_obj)
            else:
                setattr(model, getname, uploaded_file)
                model.save()
            return JsonResponse({'status': 'success'},status=200)
        else:
            return JsonResponse({"data":"error"}, status=400,safe=False)


class Groups_View(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"管理部權限管理")

    def put(self,request):
        dict_data = convent_dict(request.body)
        group = Group.objects.get(id=dict_data['id'])
        user_ids= dict_data.get("user_set",[])
        users = User.objects.filter(id__in=user_ids)
        group.user_set.set(users)
        return JsonResponse({'data': "修改成功"},status=200)


    def get(self, request):
        id = request.GET.get('id')
        data = get_object_or_404(Group, id=id)
        groups_employee = data.user_set.all()
        users = [user.id for user in groups_employee]
        data={"user_set":users,"id":data.id,"group_name":data.name}
        
        return JsonResponse({"data": data}, status=200)

class Approval_Groups_View(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"簽核流程管理")

    @staticmethod
    def del_approval(approval_target):
        name_display = approval_target.get_name_display()
        print(name_display)

        approvals_to_delete = approval_target.approvals.filter(current_status='in_process')

        senders=[] #員工

        for approval_obj in approvals_to_delete:
            employee_id = approval_obj.get_created_by.id 
            try:
                employee = Employee.objects.get(id=employee_id)
                if employee not in senders:
                    senders.append(employee)
            except Employee.DoesNotExist:
                continue

        for sender in senders:
            SysMessage.objects.create(
                Target_user=sender,
                content=f"系統調整:{name_display} 調整簽核順序，進行中的簽核已經刪除，請重新簽核"
            )

        approvals_to_delete.delete()


    def post(self,request):
        # post 處理兩個事件，【是否勾選直屬主管】以及【送出新關卡(員工)】
        dict_data = convent_dict(request.body)
        print("dict_data ",dict_data)
        approval_target = get_object_or_404(Approval_Target, id=dict_data["set_id"])
        approval_order = approval_target.approval_order

        
        if "is_checked" in dict_data:
            # 【勾選直屬主管】:
            if dict_data["is_checked"]:
                print("勾選")
                approval_order.insert(0,"x")
                approval_target.approval_order = approval_order
                approval_target.save()
            else:
            # 【取消勾選直屬主管】:
                print("取消勾選")
                while "x" in approval_order:
                    approval_order.remove("x")
                approval_target.approval_order = approval_order
                approval_target.save()

            #刪除非完成的簽核
            self.del_approval(approval_target)
            return JsonResponse({'data': "修改成功"},status=200)



        # 【送出新關卡(員工)】:
        employee_id = int(dict_data["new_stage_id"])
        approval_order.append(employee_id)
        approval_target.approval_order = approval_order
        approval_target.save()
        print("approval_target.approval_order ",approval_target.approval_order)
        
        #刪除非完成的簽核
        self.del_approval(approval_target)

        return JsonResponse({'data': "修改成功"},status=200)

    def get(self, request):
        id = request.GET.get('id')
        data = get_object_or_404(Approval_Target, id=id)
        json_data = model_to_dict(data)
        json_data["is_director"] = False
        approval_order_list = []
        
        for item in json_data['approval_order']:
            if item != "x":
                try:
                    employee = Employee.objects.get(id=item)
                    approval_order_list.append({"id": item, "name": employee.full_name})
                except Employee.DoesNotExist:
                    continue
            elif item == "x":
                json_data["is_director"] = True
        
        json_data["approval_order"]= approval_order_list
        json_data["name"]= data.get_name_display()

        return JsonResponse({"data": json_data}, status=200)
    
    def delete(self, request):
        try:
            dict_data = convent_dict(request.body)
            employee_id = int(dict_data["employee_id"])

            approval_target = get_object_or_404(Approval_Target, id=dict_data["set_id"])
            approval_order = approval_target.approval_order
            print("del: ",employee_id)
            print("Befor    : ",approval_order)
            approval_order = [x for x in approval_order if x != employee_id]
            print("after: ",approval_order)
            approval_target.approval_order = approval_order
            approval_target.save()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            print(e)
            return JsonResponse({"error":str(e)},status=500)

class Leave_Application_View(View):

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = LeaveApplicationForm(dict_data)
        if form.is_valid():
            getObject = Leave_Application.objects.get(id=dict_data['id'])
            #本人能修改
            if getObject.created_by !=request.user.employee:
                return JsonResponse({"error":"此單本人才能修改"},status=400)

            if "type_of_leave" in dict_data:
                type_of_leave = dict_data["type_of_leave"]
                del dict_data["type_of_leave"]
                getObject.type_of_leave = Leave_Param.objects.get(id=type_of_leave)
            if "substitute" in dict_data:
                substitute = dict_data["substitute"]
                del dict_data["substitute"]
                getObject.substitute = Employee.objects.get(id=substitute)
            
            getObject.update_fields_and_save(**dict_data)
            return JsonResponse({'data': "完成修改"},status=200)
        else:
            return JsonResponse({"error":form.errors},status=400)


    def post(self,request):
        form = LeaveApplicationForm(request.POST)
        if form.is_valid():
            get_leave_id = form.cleaned_data["type_of_leave"].id
            newobj =form.save(commit=False)
            leave_obj =Leave_Param.objects.get(id=get_leave_id)

             
            can_leave = leave_obj.exceeds_leave_quantity_by_year(newobj,request.user.employee)
            
            print("xccc")
            
            if can_leave:
                newobj.save()
                return JsonResponse({'data': "完成新增","id":newobj.id},status=200)
            else:
                return JsonResponse({"error":"無法請假，已沒有請假時數"},status=400)
        else:
            print("is_valid FALSE")
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)

    
    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Leave_Application, id=id)
        data = model_to_dict(data)
        data['attachment'] = data['attachment'].url if data['attachment']  else None
        return JsonResponse({"data":data}, status=200,safe = False)

    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            getObject =Leave_Application.objects.get(id=int(dict_data['id']))
            if getObject.created_by !=request.user.employee:
                return JsonResponse({"error":"此單本人才能刪除"},status=400)
            if getObject.Approval:
                getObject.Approval.delete()
            getObject.delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            print(e)
            return JsonResponse({"error":str(e)},status=500)

class Clock_Correction_Application_View(View):

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = ClockCorrectionApplicationForm(dict_data)
        if form.is_valid():
            getObject = Clock_Correction_Application.objects.get(id=dict_data['id'])
            
            if getObject.created_by !=request.user.employee:
                return JsonResponse({"error":"此單本人才能更改"},status=400)

            if getObject.Approval :
                if(getObject.Approval.current_status=="completed" or getObject.Approval.current_status=="in_process" ):            
                    return JsonResponse({"error":"簽核進行或完成禁止修改"},status=400)


            employee = Employee.objects.get(id=request.user.employee.id)
            getObject.update_fields_and_save(**dict_data)
            clock_time = time(int(getObject.end_hours_of_clock), int(getObject.end_mins_of_clock))

            getObject.create_and_update_clock(employee, getObject.date_of_clock, getObject.type_of_clock, clock_time)
            return JsonResponse({'data': "完成修改"},status=200)
        else:
            return JsonResponse({"error":form.errors},status=400)


    def post(self,request):
        form = ClockCorrectionApplicationForm(request.POST)
        if form.is_valid():
            newobj = form.save()
            employee = Employee.objects.get(id=request.user.employee.id)
            clock_time = time(newobj.end_hours_of_clock, newobj.end_mins_of_clock)
            newobj.create_and_update_clock(employee, newobj.date_of_clock, newobj.type_of_clock, clock_time)
            return JsonResponse({'data': "完成新增","id":newobj.id},status=200)
        else:
            print("is_valid FALSE")
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)

    
    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Clock_Correction_Application, id=id)
        data = model_to_dict(data)
        data['attachment'] = data['attachment'].url if data['attachment']  else None
        return JsonResponse({"data":data}, status=200,safe = False)
    
    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            getObject=Clock_Correction_Application.objects.get(id=dict_data['id'])
            if getObject.created_by !=request.user.employee:
                return JsonResponse({"error":"此單本人才能刪除"},status=400)
            if getObject.Approval:
                getObject.Approval.delete()
            getObject.delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            print(e)
            return JsonResponse({"error":str(e)},status=500)

class Work_Overtime_Application_View(View):

    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            getObject = Work_Overtime_Application.objects.get(id=dict_data['id'])
            if getObject.created_by !=request.user.employee:
                return JsonResponse({"error":"此單本人才能刪除"},status=400)
            if getObject.Approval:
                getObject.Approval.delete()
            getObject.delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            print(e)
            return JsonResponse({"error":str(e)},status=500)


    def put(self,request):
        dict_data = convent_dict(request.body) 
        form = WorkOvertimeApplicationForm(dict_data)
        if form.is_valid():
            getObject = Work_Overtime_Application.objects.get(id=dict_data['id'])
            if getObject.created_by !=request.user.employee:
                return JsonResponse({"error":"此單本人才能更改"},status=400)

            getObject.update_fields_and_save(**dict_data)
            return JsonResponse({'data': "完成修改"},status=200)
        else:
            return JsonResponse({"error":form.errors},status=400)


    def post(self,request):
        form = WorkOvertimeApplicationForm(request.POST)
        if form.is_valid():
            newobj = form.save()
            return JsonResponse({'data': "完成新增","id":newobj.id},status=200)
        else:
            print("is_valid FALSE")
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)

    
    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Work_Overtime_Application, id=id)
        data = model_to_dict(data)
        data['attachment'] = data['attachment'].url if data['attachment']  else None
        return JsonResponse({"data":data}, status=200,safe = False)


class Leave_Param_View(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"管理部管理")

    def post(self,request):
        id = request.POST.get("id")
        form = LeaveParamModelForm(request.POST)

        if form.is_valid():
            is_audit = request.POST.get('is_audit') == 'true'
            is_attachment = request.POST.get('is_attachment') == 'true'
            data = {
                'leave_name': request.POST['leave_name'],
                'leave_quantity': request.POST['leave_quantity'],
                'minimum_leave_number': request.POST['minimum_leave_number'],
                'minimum_leave_unit': request.POST['minimum_leave_unit'],
                'leave_rules': request.POST['leave_rules'],
                'is_audit': is_audit,
                'is_attachment': is_attachment,
                'deduct_percentage': request.POST['deduct_percentage']
            }
            getobj = Leave_Param.objects.get(id=id)
            getobj.update_fields_and_save(**data)    
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)



class Profile_View(View):
    # 同一個post要處理更新照片以及更新密碼
    def post(self,request):
        print("Profile_View")
        if 'profile_image' in request.FILES:
            uploaded_image = request.FILES.get('profile_image')
            employee = get_object_or_404(Employee, id=request.user.employee.id)
            print(employee.profile_image.path)
            print(employee)
            now = datetime.datetime.now()
            date_time_string = now.strftime("%Y%m%d_%H%M%S")
            file_extension = os.path.splitext(uploaded_image.name)[1]
            new_file_name = f"{date_time_string}{file_extension}"
            employee.profile_image.save(new_file_name, uploaded_image)
            print("成功")
            return JsonResponse({'status': 'success'},status=200)
        else:
            print(request.POST)
            form = PasswordChangeForm(user=request.user, data=request.POST)
            if form.is_valid():
                newobj =form.save()
                update_session_auth_hash(request, form.user)
                return JsonResponse({'status': 'success'},status=200)
            else:
                return JsonResponse({"data":form.errors}, status=400,safe=False)

class Employee_Pasword_View(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"管理部管理")

    def post(self,request):
        form = EmployeeForm(request.POST)

        if form.is_valid():
            id = request.POST.get('id')
            password = request.POST.get('password')
            if len(password)<4:
                return JsonResponse({"error":"長度至少5碼"},status=400)
            employee = Employee.objects.get(id=id)
            print(employee)
            employee.user.set_password(password)
            employee.user.save()
            return JsonResponse({'data': "完成新增","id":employee.id},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)


class Employee_View(UserPassesTestMixin,View):
    def test_func(self):
        return Check_Permissions(self.request.user,"管理部管理")

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = EmployeeForm(dict_data)
        if form.is_valid():
            getObject=Employee.objects.get(id=dict_data['id'])
            if "labor_pension" in dict_data:
                if int(dict_data['labor_pension']) <6:
                    return JsonResponse({"error":"勞退比例至少要為6%"},status=400)

            if "departments" in dict_data:
                Department_instance = Department.objects.get(pk=int(dict_data["departments"]))
                getObject.departments = Department_instance
                del dict_data["departments"]
            else:
                getObject.departments = None

            employee_id = dict_data["employee_id"]
            if getObject.user.username !=employee_id:
                if  User.objects.filter(username=employee_id).exists():
                    return JsonResponse({"error":"員工ID已存在或是曾被刪除過，請使用別的ID。"},status=400)

                getObject.user.username = employee_id
                getObject.user.save()


            getObject.update_fields_and_save(**dict_data)
            return JsonResponse({'data': "完成修改"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)


    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            empolyee = Employee.objects.get(id=dict_data['id'])
            empolyee.user.is_active =False
            empolyee.user.save()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)


    def post(self,request):
        form = EmployeeForm(request.POST)

        if form.is_valid():
            username = form.cleaned_data['employee_id']
            password = form.cleaned_data['id_number']
            if  User.objects.filter(username=username).exists():
                return JsonResponse({"error":"員工編號已存在或是曾被刪除過,請使用別的ID。"},status=400)

            user = User.objects.create_user(username=username, password=password)
            employee = form.save(commit=False)
            employee.user = user
            employee.save()
            return JsonResponse({'data': "完成新增","id":employee.id},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)


    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Employee, id=id)
        uploaded_files = data.uploaded_files.all() 
        uploaded_files_dict_list = []
        for file in uploaded_files:
            file_dict = model_to_dict(file)
            file_dict["file"] = file.file.url
            uploaded_files_dict_list.append(file_dict)

        data = model_to_dict(data)
        data["annualleaves"] = ""
        data["uploaded_files"]=uploaded_files_dict_list
        data['profile_image'] = data['profile_image'].url if data['profile_image']  else None
        print(data)
        return JsonResponse({"data":data}, status=200,safe = False)

# 員工出勤
class Employee_Attendance_View(View):
    def get(self,request):
        department = request.GET.get('department') # 回傳department的id
        full_name = request.GET.get('full_name')
        clock_inout = request.GET.get('clock_inout')
        clock_time_date = request.GET.get('clock_time_date')
        # 篩選部門以及名稱有包含
        employees = Employee.objects.filter(departments__in=[department]).filter(full_name__icontains=full_name)
        # if clock_time_date:
        #     print(clock_time_date)
            # T是簽到F是簽退
            # clock_time_date = employee.clock.filter(created_date__in=[clock_time_date])
        data = []
        for employee in employees:
            print(employee.full_name)
            # 篩選簽到還是簽退以及有無包含過來的日期
            for clock in employee.clock.filter(clock_in_or_out__in=[clock_inout]).filter(created_date__icontains=clock_time_date):
                data.append({
                    'clock_date':clock.clock_date,
                    'department': employee.departments.department_name,
                    'employee_id': employee.employee_id,
                    'full_name': employee.full_name,
                    'clock_inout': clock_inout,
                    'clock_out': clock.clock_time,
                    'clock_GPS': clock.clock_GPS,
                })
        

        return JsonResponse({"data":data}, status=200,safe = False)



class Project_Confirmation_View(UserPassesTestMixin,View):
    def test_func(self):        
        return Check_Permissions(self.request.user,"工程確認單管理",self.request.method,"工程確認單查看") 

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = ProjectConfirmationForm(dict_data)
        if form.is_valid():
            getObject = Project_Confirmation.objects.get(id=dict_data['id'])
            if "completion_report_employee" in dict_data:
                get_completion_report_employee = dict_data["completion_report_employee"]
                del dict_data["completion_report_employee"]
                getObject.completion_report_employee.set(get_completion_report_employee)


            get_Quotation_Object = Quotation.objects.get(id=dict_data['quotation'])
            getObject.quotation = get_Quotation_Object
            del dict_data["quotation"]
            getObject.update_fields_and_save(**dict_data)

            return JsonResponse({'data': "完成修改"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)

    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            Project_Confirmation.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            print(e)
            return JsonResponse({"error":str(e)},status=500)

    def post(self,request):
        form = ProjectConfirmationForm(request.POST)

        if form.is_valid():
            newobj =form.save()
            return JsonResponse({'data': "完成新增","id":newobj.id},status=200)
        else:
            print("is_valid FALSE")
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)


    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Project_Confirmation, id=id)
        get_id=data.get_show_id()
        project_name=data.quotation.project_name if data.quotation else None
        client_name=data.quotation.client.client_name if data.quotation.client else None
        requisition_name=data.quotation.requisition.client_name if data.quotation.requisition else None

        data = model_to_dict(data)
        data['project_confirmation_id'] = get_id
        data['project_name'] = project_name
        data['client_name'] = client_name
        data['requisition_name'] = requisition_name
        data["completion_report_employee"] = convent_employee(data["completion_report_employee"])
        data['attachment'] = data['attachment'].url if data['attachment']  else None
        print(data)
        return JsonResponse({"data":data}, status=200,safe = False)

# 派任單
class Job_Assign_View(UserPassesTestMixin,View):
    def test_func(self):        
        return Check_Permissions(self.request.user,"工程派任計畫管理",self.request.method,"工程派任計畫查看")  or  Check_Permissions(self.request.user,"工程確認單管理",self.request.method,"工程確認單查看") 

    
    def put(self,request):
        dict_data = convent_dict(request.body)
        form = ProjectJobAssignForm(dict_data)
        if form.is_valid():
            getObject = Project_Job_Assign.objects.get(id=int(dict_data['id']))
            process_key =("support_employee","work_employee","lead_employee","project_confirmation")

            for key in process_key:#處理特別key
                if key in dict_data:
                    if key == "project_confirmation":
                        confirmation_instance = Project_Confirmation.objects.get(pk=int(dict_data["project_confirmation"]))
                        getObject.project_confirmation = confirmation_instance
                    else:
                        field = getattr(getObject, key)
                        field.set(dict_data[key])

            for key in process_key:#刪除處理完的key
                if key in dict_data:
                    del dict_data[key]

            if "vehicle" in dict_data:
                related_project_job_assign = getObject
                vehicle_id = dict_data["vehicle"]
                vehicle_id_list = [int(item) for item in vehicle_id]
                related_project_job_assign.vehicle.set(vehicle_id_list)
                related_project_job_assign.save()
                del dict_data["vehicle"]


            getObject.update_fields_and_save(**dict_data)

            return JsonResponse({'data': "完成修改"},status=200)
        else:
            return JsonResponse({"error":form.errors},status=400)


    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            Project_Job_Assign.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            print(e)
            return JsonResponse({"error":str(e)},status=500)

    def post(self,request):
        form = ProjectJobAssignForm(request.POST)

        if form.is_valid():
            newobj =form.save()
            return JsonResponse({'data':"完成新增","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print("error_messages: ",error_messages)
            return JsonResponse({"error":error_messages},status=400)

    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Project_Job_Assign, id=id)
        show_data_id= data.get_show_id()
        #渲染關聯
        selected_fields = ['id','quotation_id', 'quotation', 'client', 'requisition']
        quotation_selected_fields = ['id','project_name',  'client']
        requisition_name = ""
        if data.project_confirmation.quotation.requisition:
            requisition_name =data.project_confirmation.quotation.requisition.client_name
        project_confirmation_dict = model_to_dict(data.project_confirmation, fields=selected_fields)
        
        quotation_dict = model_to_dict(data.project_confirmation.quotation,fields=quotation_selected_fields)
        q_id= data.project_confirmation.quotation.get_show_id()
        client_name= data.project_confirmation.quotation.client.client_name
        quotation_dict["q_id"] = q_id
        quotation_dict["client_name"] = client_name

        car_id =[]
        for car in data.vehicle.all():
            car_id.append(str(car.id))

        data = model_to_dict(data)
        data["requisition_name"]=requisition_name
        data["vehicle"]=car_id
        data["lead_employee"] = convent_employee(data["lead_employee"])
        data["work_employee"] = convent_employee(data["work_employee"])        
       

        #將外來鍵的關聯 加入dict
        data['project_confirmation'] = project_confirmation_dict
        data['quotation_dict'] = quotation_dict
        data['job_assign_id'] = show_data_id
        return JsonResponse({"data":data}, status=200,safe = False)


class Travel_Application_View(View):

    def put(self,request):
        dict_data = convent_dict(request.body)
        form = Travel_ApplicationForm(dict_data)
        if form.is_valid():
            getObject = Travel_Application.objects.get(id=int(dict_data['id']))
            if getObject.created_by !=request.user.employee:
                return JsonResponse({"error":"此單本人才能更改"},status=400)

            getObject.update_fields_and_save(**dict_data)

            return JsonResponse({'data': "完成修改"},status=200)
        else:
            return JsonResponse({"error":form.errors},status=400)


    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            getObject=Travel_Application.objects.get(id=dict_data['id'])
            if getObject.created_by !=request.user.employee:
                return JsonResponse({"error":"此單本人才能刪除"},status=400)
            if getObject.Approval:
                getObject.Approval.delete()
            getObject.delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            print(e)
            return JsonResponse({"error":str(e)},status=500)

    def post(self,request):
        form = Travel_ApplicationForm(request.POST)

        if form.is_valid():
            newobj =form.save()
            return JsonResponse({'data':"完成新增","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print("error_messages: ",error_messages)
            return JsonResponse({"error":error_messages},status=400)

    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Travel_Application, id=id)
        show_data_id= data.get_show_id()
        data = model_to_dict(data)
        data["show_data_id"]=show_data_id
        data['attachment'] = data['attachment'].url if data['attachment']  else None

        return JsonResponse({"data":data}, status=200,safe = False)




class ExcelExportView(View):
   def post(self,request):
        json_data = json.loads(request.body)
        headers = json_data['headers']
        table_data = json_data['data']
        # print(table_data) # table_data是[{表頭名:資料內容}]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        
        # 初始化每個欄位的最大長度，以表頭長度為初始值
        max_lengths = [len(header) for header in headers]
        for row_data in table_data:
            # row = [row_data.get(header, '') for header in headers]
            row = [self.replace_br_with_newline(row_data.get(header, '')) for header in headers]
             # 更新每個欄位的最大長度
            max_lengths = [max(max_length, len(str(cell))) for max_length, cell in zip(max_lengths, row)]
            ws.append(row)
        # ws.iter_cols 迭代該列的所有欄，從 1 到 headers 的長度，也就是迭代所有欄位
        for col_idx, col in enumerate(ws.iter_cols(min_col=1, max_col=len(headers))):
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 8)  # 調整列寬的公式
            column_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[column_letter].width = adjusted_width
        # for col_idx, max_length in enumerate(max_lengths):
        #     adjusted_width = (max_length + 2) * 1.2  # 調整列寬的公式
        #     column_letter = get_column_letter(col_idx + 1)
        #     ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=表格数据.xlsx'
        wb.save(response)

        return response
   def replace_br_with_newline(self, data):
        if '<br>' in data:
            return data.replace('<br>', ' ')
        return data

class Calendar_View(View):
   def get(self,request):
        all = request.GET.get('all',"")
        if all=="1":
            related_projects = Project_Job_Assign.objects.all()
        else:
            employeeid = request.user.employee
            related_projects = Project_Job_Assign.objects.filter(lead_employee__in=[employeeid])|Project_Job_Assign.objects.filter(work_employee__in=[employeeid])
        data = []
        for project in related_projects:
            if project.attendance_date is None:
                continue
            if project.location is None:
                project.location = "暫無"
            
            lead_employee_names = project.lead_employee.all().values_list('full_name', flat=True)
            merged_names = ', '.join(lead_employee_names)
            work_employee_names = project.work_employee.all().values_list('full_name', flat=True)
            work_employee_merged_names = ', '.join(work_employee_names)

            work_method_str = "非派工"
            if project.work_method:
               work_method_str = "派工"


            vehicle = project.vehicle.all().values_list('vehicle_id', flat=True)
            merged_vehicle_id = ', '.join(vehicle)

            project_id = f"工派-{str(project.id).zfill(5)}"
            print("project_id",project_id)
            # client = project.project_confirmation.quotation.client.client_name


            quotation_obj = project.project_confirmation.quotation
            client = "" if quotation_obj.client is None else quotation_obj.client.client_name
            requisition = "" if quotation_obj.requisition is None else quotation_obj.requisition.client_name
            # requisition = project.project_confirmation.quotation.requisition.client_name

            data.append({
                'title': project.project_confirmation.quotation.project_name,
                'start': project.attendance_date,
                'work_method': work_method_str,
                'location': project.location,
                'lead_employee': merged_names,
                'work_employee': work_employee_merged_names,
                'employee_assign_id': project_id,
                'vehicle': merged_vehicle_id,
                'remark': project.remark,
                'client': client,
                'requisition': requisition
            })
            print(data)
        return JsonResponse(data, status=200,safe = False)

class Vehicle_View(View):
    def get(self,request):
        id = request.GET.get('id')
        data = get_object_or_404(Vehicle, id=id)
        data = model_to_dict(data)
        return JsonResponse({"data":data}, status=200,safe = False)
    
    def put(self,request):
        dict_data = convent_dict(request.body)
        form = VehicleForm(dict_data)
        if form.is_valid():
            Vehicle.objects.get(id=dict_data['id']).update_fields_and_save(**dict_data)
            return JsonResponse({'data': "修改成功"},status=200)
        else:
            error_messages = form.get_error_messages()
            return JsonResponse({"error":error_messages},status=400)

    def delete(self,request):
        try:
            dict_data = convent_dict(request.body)
            Vehicle.objects.get(id=dict_data['id']).delete()
            return HttpResponse("成功刪除",status=200)
        except ObjectDoesNotExist:
            return JsonResponse({"error":"資料不存在"},status=400)
        except  Exception as e:
            return JsonResponse({"error":str(e)},status=500)

    def post(self,request):
        form = VehicleForm(request.POST)
        if form.is_valid():
            newobj = form.save()
            return JsonResponse({"data":"新增成功","id":newobj.id},status=200)
        else:
            error_messages = form.get_error_messages()
            print(error_messages)
            return JsonResponse({"error":error_messages},status=400)

class Check(View):
    def post(self,request):
        data = json.loads(request.body)
        gps = data.get('gps')
        clock_in_or_out = data.get('clock_in_or_out')
        clock_time = datetime.datetime.now()

        Clock.objects.create(
            employee_id=request.user.employee,
            clock_time=clock_time,
            clock_in_or_out=clock_in_or_out,
            clock_GPS=gps
        )

        return JsonResponse({'status': 'success'},status=200)

    def get(self,request):
       return HttpResponseNotAllowed(['only POST'])



class DeleteUploadedFileView(View):
    def delete(self, request,model, obj_id, file_id):
        match model:
            case "employee":
                model=Employee.objects.get(id=obj_id)
            case "Quotation":
                model=Quotation.objects.get(id=obj_id)
            case "project_employee_assign":
                model=Project_Employee_Assign.objects.get(id=obj_id)

        uploaded_file = get_object_or_404(UploadedFile, id=file_id)
        if uploaded_file in model.uploaded_files.all():
            model.uploaded_files.remove(uploaded_file)
            model.save()
            uploaded_file.delete()            
            return JsonResponse({"message": "檔案已刪除。"}, status=200)
        else:
            return JsonResponse({"message": "檔案不存在或不屬於此員工。"}, status=400)
        



#啟動SERVER 同時 開另一個CMD  除錯:python manage.py process_tasks
#改下面的code，runserver 可能要重啟
from background_task import background
from datetime import timedelta
def calculate_annual_leave(employee):
    from datetime import datetime

    start_work_date = employee.start_work_date
    if not start_work_date:
        return
    
    years = employee.seniority()
    today = datetime.today().date()

    #計算入職的前一天的明年時間
    end_date = employee.start_work_date - timedelta(days=1)
    end_date = end_date.replace(year=today.year + 1)
    if years == "人資單位未填寫入值日" :
        return "人資單位未填寫入值日"

    if years <0.6 :
        return "小於0.6"
    
    if years < 1:
        has_three_days_leave = employee.annualleaves.filter(days=3).exists()
        if not has_three_days_leave:
            annual_leave = AnnualLeave.objects.create(days=3, end_date=end_date, remark="")
            employee.annualleaves.add(annual_leave)
        return "0.6"
    
    #今天=入職 才執行給假
    if  not (today.month == start_work_date.month and today.day == start_work_date.day):
        return "不執行"
    
    give_day = 0
    if years >= 1 and years < 2:
        give_day= 7
    elif years >= 2 and years < 3:
        give_day= 10
    elif years >= 3 and years < 5:
        give_day= 14
    elif years >= 5 and years < 10:
        give_day= 15
    elif years == 10:
        give_day= 16
    else:
        give_day= min(30, math.floor(years))

    matching_leaves = AnnualLeave.objects.filter(end_date=end_date)

    if not matching_leaves.exists():#不存在才給
        annual_leave = AnnualLeave.objects.create(days=give_day, end_date=end_date, remark="")
        employee.annualleaves.add(annual_leave)
    
    return f"年資:{years}、給假{give_day} 給過?{matching_leaves}"

#一天86400 
@background(schedule=86400)
def calculate_annual_leave_for_all_employees():
    print("go task")
    employees = Employee.objects.all()
    for employee in employees:
        msg = calculate_annual_leave(employee)
        print(employee.full_name)
        print(msg)
    print("end")
#啟動
calculate_annual_leave_for_all_employees()